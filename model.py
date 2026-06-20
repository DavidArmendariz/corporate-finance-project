"""Parametrized model of "VPN del Proyecto" from Taller_final.xlsx.

This is a pure-Python reimplementation of the cash-flow chain behind
`Estado de resultados!C68`, exposing the decision variables the user wants to
optimize as `Levers`. All other inputs are read once from the workbook as
`FixedParams` so the model stays in sync with the spreadsheet.

The model reproduces Excel's cached VPN exactly (see `validate()` at the bottom).

Non-obvious behavior replicated deliberately: `%D` (porcentaje de deuda) affects
VPN through TWO channels:
  - the WACC discount rate, and
  - `Impuestos por pagar` (row 37 = row 29), which is computed on EBT
    (= EBIT - Gastos financieros). Gastos financieros come from the Bonos
    corporate-bond schedule sized by `%D`, so debt flows into the FCL via KTNO.
NOPAT itself uses the *unlevered* operative tax (row 50). The model keeps this
mix of levered/unlevered tax exactly as the sheet does.

`Gastos por seguro` (row 25) is a constant per-year EBIT expense
(= capex_total * pct_seguro / 5, the insurance premium amortized over 5 years
sourced from `Amortización gasto seguro`). Unlike depreciation/amortization it
is NOT added back in the cash flow (it is a cash cost). `Amortización` (row 23)
is now zero. Both `Comisiones` and `Porcentaje de seguro` are decision levers.
"""

from __future__ import annotations

from dataclasses import dataclass, field

import openpyxl

WORKBOOK_PATH = "Taller_final.xlsx"
N_YEARS = 10  # Año 1 .. Año 10 (columns D..M in the sheet)


@dataclass
class FixedParams:
    """Inputs to VPN held constant during optimization (read from the workbook)."""

    inflation: list[float]          # ES!D4:M4, per year (Año 1..10)
    tax: float                      # Datos!C17 (== ES D5:M5 == WACC!B6)
    precio_base: float              # ES!D8 (hardcoded =6500000)
    g_real_mp: float                # Datos!C23, real growth of raw-material cost
    nomina_oper_base: float         # Nómina!H3 (grows by inflation)
    nomina_admin_base: float        # Nómina!H4 (grows by inflation)
    pct_mantenimiento: float        # Datos!C21 (of revenue)
    pct_gastos_op_var: float        # Datos!C26 (of revenue)
    depre_annual: float             # constant per year (Depreciaciones!B7)
    capex_total_mm: float           # Datos!C10 (millones de COP)
    pct_overhaul: float             # Datos!C24 (year-5 renovation, of initial capex)
    multa_base: float               # Datos!C19 * 1e6 (Costo de oportunidad!B18 penalty)
    # Salvage (Año 11) constants
    pct_venta_activos: float        # ES!P37
    pct_desmantelamiento: float     # ES!P38
    tax_ganancia_ocasional: float   # ES!P39
    # WACC building blocks
    rf: float                       # WACC!B1
    market_return: float            # WACC!B2
    beta_l: float                   # WACC!B4
    de_sector: float                # WACC!B5 (D/E of the sector)
    bono_yankee: float              # WACC!B11
    tes_cop: float                  # WACC!B13
    # Bonos (corporate bond — cost of debt and the base annual schedule)
    kd_cop: float                   # WACC!B17 == Bonos!D17, effective annual cost of debt
    pct_d_base: float               # Datos!C35, the %D the bond series below were computed at
    bond_cupones_mm: list[float]    # Bonos!G27:G36, annual coupons (millones de COP)
    bond_servicio_mm: list[float]   # Bonos!I27:I36, annual pago total (millones de COP)
    bond_deuda_mm: list[float]      # Bonos!J27:J36, year-end balance (millones de COP)

    @property
    def capex_total(self) -> float:
        return self.capex_total_mm * 1_000_000

    @property
    def salvage(self) -> float:
        """ES!P49 — Año-11 cash flow from asset disposal. Independent of levers."""
        valor_venta = self.capex_total * self.pct_venta_activos          # P42
        ppe_neta_10 = self.capex_total - N_YEARS * self.depre_annual     # M43
        utilidad = valor_venta - ppe_neta_10                             # P44
        impuesto = max(0.0, utilidad * self.tax_ganancia_ocasional)      # P45
        flujo_venta = valor_venta - impuesto                            # P47
        desmantelamiento = self.capex_total_mm * self.pct_desmantelamiento * 1_000_000  # P48
        return flujo_venta - desmantelamiento                           # P49


@dataclass
class Levers:
    """The decision variables to be optimized."""

    pct_e: float                    # Datos!C34, porcentaje de equity
    pct_d: float                    # Datos!C35, porcentaje de deuda
    costo_mp_base: float            # ES!D9, raw-material cost per ton, Año 1
    cxc_dias: float                 # Datos!C28
    cxp_dias: float                 # Datos!C29
    g_real_precio: float            # Datos!C31, real annual price growth
    pct_comisiones: float           # Datos!C27 (of revenue)
    pct_seguro: float               # Datos!C25 (insurance, of initial capex)
    cantidades: list[float] = field(default_factory=list)  # ES!D10:M10 (10 values)


def load_fixed_params(path: str = WORKBOOK_PATH) -> tuple[FixedParams, Levers]:
    """Read constants (FixedParams) and the base-case lever values from the workbook."""
    wb = openpyxl.load_workbook(path, data_only=True)
    es, d, w, bonos, nom = (
        wb["Estado de resultados"],
        wb["Datos"],
        wb["WACC"],
        wb["Bonos"],
        wb["Nómina"],
    )

    inflation = [es.cell(row=4, column=col).value for col in range(4, 4 + N_YEARS)]  # D4:M4
    depre_annual = sum(
        d.cell(row=r, column=3).value * 1_000_000 / d.cell(row=r, column=5).value
        for r in range(5, 10)  # Datos C5:C9 / E5:E9
    )

    fixed = FixedParams(
        inflation=inflation,
        tax=d["C17"].value,
        precio_base=es["D8"].value,
        g_real_mp=d["C23"].value,
        nomina_oper_base=nom["H3"].value,
        nomina_admin_base=nom["H4"].value,
        pct_mantenimiento=d["C21"].value,
        pct_gastos_op_var=d["C26"].value,
        depre_annual=depre_annual,
        capex_total_mm=d["C10"].value,
        pct_overhaul=d["C24"].value,
        multa_base=d["C19"].value * 1_000_000,
        pct_venta_activos=es["P38"].value,
        pct_desmantelamiento=es["P39"].value,
        tax_ganancia_ocasional=es["P40"].value,
        rf=w["B1"].value,
        market_return=w["B2"].value,
        beta_l=w["B4"].value,
        de_sector=w["B5"].value,
        bono_yankee=w["B11"].value,
        tes_cop=w["B13"].value,
        kd_cop=w["B17"].value,
        pct_d_base=d["C35"].value,
        bond_cupones_mm=[bonos.cell(row=r, column=7).value for r in range(27, 37)],   # G27:G36
        bond_servicio_mm=[bonos.cell(row=r, column=9).value for r in range(27, 37)],  # I27:I36
        bond_deuda_mm=[bonos.cell(row=r, column=10).value for r in range(27, 37)],    # J27:J36
    )

    base_levers = Levers(
        pct_e=d["C34"].value,
        pct_d=d["C35"].value,
        costo_mp_base=es["D9"].value,
        cxc_dias=d["C28"].value,
        cxp_dias=d["C29"].value,
        g_real_precio=d["C31"].value,
        pct_comisiones=d["C27"].value,
        pct_seguro=d["C25"].value,
        cantidades=[es.cell(row=10, column=col).value for col in range(4, 4 + N_YEARS)],
    )
    return fixed, base_levers


@dataclass
class NominaRole:
    """One row of the Nómina roster (Nómina!A3:E17)."""

    cargo: str            # Nómina col A
    count: int            # Nómina col B (No)
    salario_mes: float    # Nómina col C (Salario / Mes, millones de COP)
    es_produccion: bool   # Nómina col E ("SI" -> operativa, "NO" -> administrativa)


def load_nomina_roster(path: str = WORKBOOK_PATH) -> list[NominaRole]:
    """Read the 15 nómina roles (Nómina!A3:E17) feeding H3/H4."""
    wb = openpyxl.load_workbook(path, data_only=True)
    nom = wb["Nómina"]
    roster = []
    for r in range(3, 18):  # rows 3..17
        cargo = nom.cell(row=r, column=1).value
        count = nom.cell(row=r, column=2).value
        salario_mes = nom.cell(row=r, column=3).value
        flag = nom.cell(row=r, column=5).value
        roster.append(
            NominaRole(
                cargo=str(cargo),
                count=int(count),
                salario_mes=float(salario_mes),
                es_produccion=str(flag).strip().upper() == "SI",
            )
        )
    return roster


def nomina_bases(roster: list[NominaRole]) -> tuple[float, float]:
    """Return (operativa, administrativa) annual bases in COP, mirroring Nómina!H3/H4.

    Annual salary per role = count * salario_mes * 12 (millones), summed by SI/NO and
    scaled to COP. SI roles feed nómina operativa; NO roles feed nómina administrativa.
    """
    oper = sum(r.count * r.salario_mes * 12 for r in roster if r.es_produccion) * 1e6
    admin = sum(r.count * r.salario_mes * 12 for r in roster if not r.es_produccion) * 1e6
    return oper, admin


def debt_schedule_interest(pct_d: float, fixed: FixedParams) -> list[float]:
    """Return the 10 annual bond coupons (Bonos!G27:G36, in millones de COP), scaled to %D.

    The Bonos sheet computes a semi-annual, IPC-indexed coupon schedule for a bullet bond
    (principal repaid in full at Año 10). Its annual summary is cached in the workbook at
    the base `%D = Datos!C35`. The bond monto = Datos!C10 * %D, and every coupon is
    `saldo * tasa_periodica` with `saldo` constant (bullet), so coupons scale LINEARLY with
    `%D`. We read the base series once and rescale, which is exact and avoids replicating
    the sheet's TODAY()-based day-count (which would drift by calendar date).

    The result feeds `ES!D27 = TRANSPOSE(Bonos!G27:G36)*1000000` (see compute_model for the
    ×1e6 scaling), which reproduces Excel's VPN exactly.
    """
    scale = pct_d / fixed.pct_d_base
    return [c * scale for c in fixed.bond_cupones_mm]


def debt_schedule_annual(pct_d: float, fixed: FixedParams) -> tuple[list[float], list[float]]:
    """Return (servicio_deuda, deuda) per year, in COP — Bonos RESUMEN ANUAL I27:I36 / J27:J36.

    `servicio` (Pago Total, col I) = annual coupons + principal amortization (the bullet
    repayment lands in Año 10). `deuda` (col J) = year-end balance (the issue amount through
    Año 9, then 0). Both scale linearly with `%D` (bond monto = Datos!C10 * %D), so we read
    the base series and rescale, then convert millones de COP to COP (×1e6). They feed ES
    rows 87 (servicio) and 89 (deuda).
    """
    scale = pct_d / fixed.pct_d_base
    servicio = [s * scale * 1_000_000 for s in fixed.bond_servicio_mm]
    deuda = [d * scale * 1_000_000 for d in fixed.bond_deuda_mm]
    return servicio, deuda


# Per-year indicators graphed in the app (order matches the dashboard layout).
# kind: "pct" | "millones" | "x"  ·  chart: "line" | "bar"
METRIC_SPECS = [
    {"key": "eva", "title": "EVA", "kind": "millones", "chart": "line"},
    {"key": "roic", "title": "ROIC", "kind": "pct", "chart": "line"},
    {"key": "ebitda", "title": "EBITDA", "kind": "millones", "chart": "line"},
    {"key": "margen_bruto", "title": "Margen bruto", "kind": "pct", "chart": "bar"},
    {"key": "ktno", "title": "KTNO", "kind": "millones", "chart": "line"},
    {"key": "palanca", "title": "Palanca de crecimiento", "kind": "x", "chart": "bar"},
    {"key": "margen_ebitda", "title": "Margen EBITDA", "kind": "pct", "chart": "bar"},
    {"key": "margen_ebit", "title": "Margen EBIT", "kind": "pct", "chart": "bar"},
    {"key": "margen_ktno", "title": "Productividad del capital de trabajo",
     "kind": "pct", "chart": "bar"},
    {"key": "cobertura_intereses", "title": "Razón de cobertura de intereses",
     "kind": "x", "chart": "bar"},
    {"key": "dscr", "title": "Razón de cobertura de servicio a la deuda",
     "kind": "x", "chart": "bar"},
    {"key": "apalancamiento", "title": "Apalancamiento financiero", "kind": "x", "chart": "bar"},
]


def wacc(pct_e: float, pct_d: float, fixed: FixedParams) -> float:
    """Reproduce WACC!B18 = Ke_COP*%E + %D*(1-tax)*Kd."""
    tax = fixed.tax
    erp = fixed.market_return - fixed.rf                          # B3
    beta_u = fixed.beta_l / (1 + (1 - tax) * fixed.de_sector)     # B7
    de_ratio = pct_d / pct_e                                      # B9
    beta_proyecto = beta_u * (1 + (1 - tax) * de_ratio)          # B10
    rp = fixed.bono_yankee - fixed.rf                            # B12
    deval = fixed.tes_cop - fixed.bono_yankee                    # B14
    ke_usd = fixed.rf + beta_proyecto * erp + rp                 # B15
    ke_cop = (1 + ke_usd) * (1 + deval) - 1                      # B16
    return ke_cop * pct_e + pct_d * (1 - tax) * fixed.kd_cop      # B18


@dataclass
class ModelResult:
    vpn: float
    wacc: float
    fcl_unlevered: list[float]  # C64 .. N64 (years 0..11), 12 values
    salvage: float
    metrics: dict[str, list[float]]  # per-year indicators (keys in METRIC_SPECS), 10 values each


def compute_model(levers: Levers, fixed: FixedParams) -> ModelResult:
    """Build the full 10-year P&L + cash-flow chain and return VPN and key intermediates."""
    n = N_YEARS
    infl = fixed.inflation
    tax = fixed.tax
    cant = levers.cantidades
    assert len(cant) == n, f"cantidades must have {n} values"

    # --- Precio (row 8) and Costo (row 9): grow by inflation_of_previous_year * (1+real) ---
    precio = [0.0] * n
    costo = [0.0] * n
    precio[0] = fixed.precio_base
    costo[0] = levers.costo_mp_base
    for j in range(1, n):
        precio[j] = precio[j - 1] * (1 + infl[j - 1]) * (1 + levers.g_real_precio)
        costo[j] = costo[j - 1] * (1 + infl[j - 1]) * (1 + fixed.g_real_mp)

    ingresos = [precio[j] * cant[j] for j in range(n)]        # row 15
    costo_ventas = [costo[j] * cant[j] for j in range(n)]     # row 16

    # Nómina (rows 17, 20) grow by inflation of previous year
    nom_oper = [0.0] * n
    nom_admin = [0.0] * n
    nom_oper[0] = fixed.nomina_oper_base
    nom_admin[0] = fixed.nomina_admin_base
    for j in range(1, n):
        nom_oper[j] = nom_oper[j - 1] * (1 + infl[j - 1])
        nom_admin[j] = nom_admin[j - 1] * (1 + infl[j - 1])

    mantenimiento = [ing * fixed.pct_mantenimiento for ing in ingresos]   # row 18
    utilidad_bruta = [
        ingresos[j] - costo_ventas[j] - nom_oper[j] - mantenimiento[j] for j in range(n)
    ]  # row 19
    gastos_op_var = [ing * fixed.pct_gastos_op_var for ing in ingresos]   # row 21
    comisiones = [ing * levers.pct_comisiones for ing in ingresos]        # row 24
    depre = fixed.depre_annual
    amort = 0.0                                                           # row 23 (== ES!D23)
    # Gastos por seguro (row 25): insurance premium amortized over 5 years (280M/yr).
    # Quirk replicated deliberately: only the Año-1 EBIT formula subtracts it
    # (D26 = D19-...-D25); E26:M26 omit the seguro term, so the workbook deducts
    # insurance in Año 1 only. seguro[j] = expense for j==0, else 0.
    gasto_seguro = fixed.capex_total_mm * levers.pct_seguro * 1_000_000 / 5
    seguro = [gasto_seguro] + [0.0] * (n - 1)

    ebit = [
        utilidad_bruta[j] - nom_admin[j] - gastos_op_var[j]
        - depre - amort - comisiones[j] - seguro[j]
        for j in range(n)
    ]  # row 26 (D26 includes -D25 seguro; E26:M26 omit it)
    impuestos_op = [max(0.0, ebit[j] * tax) for j in range(n)]            # row 50 (unlevered)
    nopat = [ebit[j] - impuestos_op[j] for j in range(n)]                 # row 51
    flujo_caja_bruto = [nopat[j] + depre + amort for j in range(n)]       # rows 54/60

    # --- Levered branch feeding KTNO via Impuestos por pagar ---
    interest_mm = debt_schedule_interest(levers.pct_d, fixed)
    # ES!D27 = TRANSPOSE(Bonos!G27:G36)*1000000 (cupones are in millones de COP).
    gastos_fin = [interest_mm[j] * 1_000_000 for j in range(n)]           # row 27
    ebt = [ebit[j] - gastos_fin[j] for j in range(n)]                    # row 28
    impuestos_por_pagar = [max(0.0, ebt[j] * tax) for j in range(n)]     # rows 29/37

    cxc = [(levers.cxc_dias / 360) * ingresos[j] for j in range(n)]      # row 34 (= KTO row 35)
    cxp = [(levers.cxp_dias / 360) * costo_ventas[j] for j in range(n)]  # row 36
    ktno = [cxc[j] - cxp[j] - impuestos_por_pagar[j] for j in range(n)]  # row 38

    d_ktno = [0.0] * n  # row 39
    d_ktno[0] = ktno[0]
    for j in range(1, n):
        d_ktno[j] = ktno[j] - ktno[j - 1]

    fcl = [flujo_caja_bruto[j] - d_ktno[j] for j in range(n)]            # row 62

    # CAPEX (row 63): year0 = -total; year5 (Año 5 = index 4) = overhaul; else 0
    capex = [0.0] * n
    capex[4] = -fixed.capex_total * fixed.pct_overhaul
    capex0 = -fixed.capex_total

    # FCL unlevered (row 64): years 0..10 then salvage as year 11
    fcl_unlevered = [capex0] + [fcl[j] + capex[j] for j in range(n)] + [fixed.salvage]

    # VPN (C71): NPV of years 1..11 at WACC, plus undiscounted year 0
    w = wacc(levers.pct_e, levers.pct_d, fixed)
    npv = sum(fcl_unlevered[k] / (1 + w) ** k for k in range(1, len(fcl_unlevered)))
    vpn = npv + fcl_unlevered[0]

    # --- Per-year financial indicators (ES rows 32, 38, 77-90) ---
    def _div(a: float, b: float) -> float:
        return a / b if b else float("nan")

    ebitda = [ebit[j] + depre + amort for j in range(n)]                 # row 32
    ppe_neta = [fixed.capex_total - depre * (j + 1) for j in range(n)]   # row 43
    capital_invertido = [ppe_neta[j] + ktno[j] for j in range(n)]        # row 45
    roic = [_div(nopat[j], capital_invertido[j]) for j in range(n)]      # row 77
    eva = [(roic[j] - w) * capital_invertido[j] for j in range(n)]       # row 84
    margen_bruto = [_div(utilidad_bruta[j], ingresos[j]) for j in range(n)]  # row 78
    margen_ebit = [_div(ebit[j], ingresos[j]) for j in range(n)]         # row 79
    margen_ebitda = [_div(ebitda[j], ingresos[j]) for j in range(n)]     # row 80
    margen_ktno = [_div(ktno[j], ingresos[j]) for j in range(n)]         # row 81
    palanca = [_div(margen_ebitda[j], margen_ktno[j]) for j in range(n)]  # row 83
    cobertura_intereses = [_div(ebit[j], gastos_fin[j]) for j in range(n)]  # row 86
    servicio, deuda = debt_schedule_annual(levers.pct_d, fixed)          # rows 87, 89
    dscr = [_div(fcl[j], servicio[j]) for j in range(n)]                 # row 88
    apalancamiento = [_div(deuda[j], ebitda[j]) for j in range(n)]       # row 90

    metrics = {
        "eva": eva,
        "roic": roic,
        "ebitda": ebitda,
        "margen_bruto": margen_bruto,
        "ktno": ktno,
        "palanca": palanca,
        "margen_ebitda": margen_ebitda,
        "margen_ebit": margen_ebit,
        "margen_ktno": margen_ktno,
        "cobertura_intereses": cobertura_intereses,
        "dscr": dscr,
        "apalancamiento": apalancamiento,
    }

    return ModelResult(
        vpn=vpn, wacc=w, fcl_unlevered=fcl_unlevered, salvage=fixed.salvage, metrics=metrics
    )


def compute_vpn(levers: Levers, fixed: FixedParams) -> float:
    """Convenience wrapper: return only the VPN (objective for the solver)."""
    return compute_model(levers, fixed).vpn


# --- Cached Excel values used to validate the reimplementation ---
EXCEL_VPN = -2_987_859_420.2771225  # ES!C68
EXCEL_WACC = 0.18189262150651486    # WACC!B18
EXCEL_FCL64 = [                     # ES!C64:N64 (years 0..11)
    -14_000_000_000,
    -581_333_333.3333333,
    842_844_525.3333333,
    1_040_281_227.2254083,
    2_377_727_442.2270985,
    -6_564_482_256.972959,
    3_815_583_986.8138485,
    8_293_798_510.254985,
    9_735_704_794.08979,
    11_239_395_421.421486,
    12_949_635_299.331955,
    2_730_000_000,
]
EXCEL_SALVAGE = 2_730_000_000.0


def validate() -> None:
    fixed, base = load_fixed_params()
    res = compute_model(base, fixed)

    print("=== Validation against Excel cached values ===")
    print(f"WACC     model={res.wacc:.15f}  excel={EXCEL_WACC:.15f}")
    print(f"Salvage  model={res.salvage:,.2f}  excel={EXCEL_SALVAGE:,.2f}")
    print(f"VPN      model={res.vpn:,.4f}")
    print(f"VPN      excel={EXCEL_VPN:,.4f}")
    print(f"VPN diff       {res.vpn - EXCEL_VPN:,.6f}")

    print("\nFCL unlevered (row 64), model vs excel:")
    for k, (m, e) in enumerate(zip(res.fcl_unlevered, EXCEL_FCL64)):
        flag = "OK" if abs(m - e) < 1.0 else "MISMATCH"
        print(f"  year {k:>2}: model={m:>22,.2f}  excel={e:>22,.2f}  [{flag}]")

    assert abs(res.wacc - EXCEL_WACC) < 1e-12, "WACC mismatch"
    assert abs(res.salvage - EXCEL_SALVAGE) < 1.0, "Salvage mismatch"
    for m, e in zip(res.fcl_unlevered, EXCEL_FCL64):
        assert abs(m - e) < 1.0, "FCL stream mismatch"
    assert abs(res.vpn - EXCEL_VPN) < 1.0, "VPN mismatch"
    print("\nAll checks passed: model reproduces Excel's VPN.")


def sensitivity_check() -> None:
    """Print VPN deltas when each lever moves in its value-improving direction."""
    from dataclasses import replace

    fixed, base = load_fixed_params()
    base_vpn = compute_vpn(base, fixed)
    print("\n=== Sensitivity (each lever nudged toward higher VPN) ===")
    print(f"base VPN = {base_vpn:,.2f}")

    nudges = {
        "costo_mp_base -10%": replace(base, costo_mp_base=base.costo_mp_base * 0.9),
        "cxc_dias -10 days": replace(base, cxc_dias=base.cxc_dias - 10),
        "cxp_dias +10 days": replace(base, cxp_dias=base.cxp_dias + 10),
        "g_real_precio +1pp": replace(base, g_real_precio=base.g_real_precio + 0.01),
        "comisiones -1pp": replace(base, pct_comisiones=base.pct_comisiones - 0.01),
        "seguro -1pp": replace(base, pct_seguro=base.pct_seguro - 0.01),
        "cantidades +10%": replace(base, cantidades=[q * 1.1 for q in base.cantidades]),
        "pct_d -> 0.2 (pct_e 0.8)": replace(base, pct_d=0.2, pct_e=0.8),
    }
    for name, lev in nudges.items():
        v = compute_vpn(lev, fixed)
        print(f"  {name:<28} VPN = {v:>22,.2f}   Δ = {v - base_vpn:>20,.2f}")


if __name__ == "__main__":
    validate()
    sensitivity_check()
